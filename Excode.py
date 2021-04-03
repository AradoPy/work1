from openpyxl import load_workbook
import geojson as jsn
import random as rd
from pymongo import MongoClient
from bson.json_util import dumps, loads
import mongo_to_geojson
import pprint

# Читает файл xlsx и выводит список координат.
def reader():
    wb = load_workbook('./data.xlsx')
    sheet = wb['Лист1']
    x = []
    y = []
    for i in range(1, 129):
            a = sheet.cell(row=i+1, column=2).value
            x.append(float(a))
    for i in range(1, 129):
            a = sheet.cell(row=i+1, column=3).value
            y.append(float(a))
    coor = zip(y, x)
    list_coor = list(coor)
    return list_coor

# Формирует geojson и json для чтения базы данных
def list2geojson(list):
    features = []
    for i in list:
        # Создание случайных аргументов для параметров gejson
        tupe_road = ["urb", "rur", "hway"]
        width_road = [5, 7, 10, 12, 18]
        long_road = rd.randint(100, 150)
        surface = ["asp", "crt", "gra", "grnd"]
        # Создаём фичи geojson
        features.append(jsn.Feature(geometry=jsn.Point(i), properties={"type": rd.choice(tupe_road),
                                                                       "width": rd.choice(width_road),
                                                                       "long": long_road,
                                                                       "face": rd.choice(surface),
                                                                       }))
    # Создаём коллекцию фич
    feature_collection = jsn.FeatureCollection(features=features)
    with open('All_points.geojson', 'w') as f:
        jsn.dump(feature_collection, f)
    with open('JSON_for_DB.json', 'w') as f:
        jsn.dump(features, f)

# Подгрузка файла в базу MongoDB
# Обращаться только 1 раз для загрузки
def load2mongodb():
    client = MongoClient('localhost', 27017)
    db = client['MyDB']
    collection_currency = db['MyColl']

    with open('JSON_for_DB.json') as f:
        file_data = jsn.load(f)
    collection_currency.insert_many(file_data)
    client.close()

# Осуществляет подключение и формирует выборку из базы данных
def sample(property, type):
    client = MongoClient('localhost', 27017)
    db = client['MyDB']
    collection_currency = db['MyColl']
    # Создаём пустые фичи
    new_features = []
    for i in collection_currency.find({property: type}):
        new_features.append(jsn.Feature(geometry=i['geometry'], properties=i['properties']))
    new_feature_collection = jsn.FeatureCollection(features=new_features)
    # Генерируем файл выборки с именем Sample + random
    with open('Sample'+str(rd.choice(range(1000,10000000))), 'w') as r:
        jsn.dump(new_feature_collection, r)

list = reader()
list2geojson(list)
# load2mongodb()
sample('properties.type','rur')
sample('properties.type','hway')
sample('properties.long', 150)
sample('properties.long', 100)
sample('properties.face', 'crt')
sample('properties.face', 'gra')
sample('properties.width', 12)
sample('properties.face', 'asp')
sample('properties.long', 125)
sample('properties.type', 'urb')
